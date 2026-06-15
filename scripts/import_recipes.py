#!/usr/bin/env python3
#
# Pi-Menu - Weekly Meal Planner
# Creator: nobody174 (nobodylearn174@gmail.com)
# GitHub: https://github.com/nobody174/Pi-Menu-Public
# License: MIT
#

"""
Enhanced Excel Recipe Importer

Reads recipes from the Excel template and converts them to JSON format
for use by Pi-Menu.

Usage:
    python3 import_recipes.py <path-to-excel-file> [--output <output-path>] [--merge]

Examples:
    python3 import_recipes.py templates/my_recipes.xlsx
    python3 import_recipes.py my_recipes.xlsx --output data/my_recipes.json
    python3 import_recipes.py recipes.xlsx --merge  # Merge with existing recipes
"""

import json
import sys
import uuid
import logging
from pathlib import Path
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)

def generate_recipe_id():
    """Generate unique recipe ID"""
    return str(uuid.uuid4())[:8]

def import_recipes(excel_path: str, output_path: str = None, merge: bool = False):
    """
    Import recipes from Excel template to JSON format.

    Args:
        excel_path: Path to the Excel template file
        output_path: Optional custom output path (defaults to data/recipes.json)
        merge: If True, merge with existing recipes; if False, replace

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl is required. Install with: pip install openpyxl")
            return False

        excel_file = Path(excel_path)
        if not excel_file.exists():
            logger.error(f"File not found: {excel_path}")
            return False

        logger.info(f"Reading recipes from: {excel_file.name}")

        wb = load_workbook(excel_file)
        if 'Add Your Recipes' not in wb.sheetnames:
            logger.error("Sheet 'Add Your Recipes' not found in workbook")
            return False

        ws = wb['Add Your Recipes']
        recipes = []
        skipped = 0

        # Process each row (skip header row 1)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue

            try:
                # Expected columns: [Name_NO, Name_EN, Category, Time, Difficulty, Servings, Allergens, ...]
                if len(row) < 7:
                    logger.warning(f"Row {row_idx}: Skipped (insufficient columns)")
                    skipped += 1
                    continue

                name_no, name_en, category, time_min, difficulty, servings, allergens = row[:7]

                if not name_no or not name_en:
                    logger.warning(f"Row {row_idx}: Skipped (missing recipe names)")
                    skipped += 1
                    continue

                recipe = {
                    "id": generate_recipe_id(),
                    "title_no": str(name_no).strip(),
                    "title_en": str(name_en).strip(),
                    "category": str(category).strip() if category else "Annet",
                    "time_minutes": int(time_min) if isinstance(time_min, (int, float)) else 30,
                    "difficulty": str(difficulty).strip() if difficulty else "Enkel",
                    "servings": int(servings) if isinstance(servings, (int, float)) else 4,
                    "allergens": [a.strip() for a in str(allergens).split(",") if a.strip()] if allergens else [],
                    "imported_at": datetime.now().isoformat(),
                    "ingredients_included": [],
                    "instructions_no": "",
                    "instructions_en": ""
                }

                recipes.append(recipe)

            except Exception as e:
                logger.warning(f"Row {row_idx}: Skipped ({str(e)})")
                skipped += 1
                continue

        # Set output path
        if output_path is None:
            output_path = Path(__file__).parent.parent / "data" / "recipes.json"
        else:
            output_path = Path(output_path)

        # Handle merge
        if merge and output_path.exists():
            with open(output_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
            recipes = existing + recipes
            logger.info(f"Merged with {len(existing)} existing recipes")

        # Save recipes
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(recipes, f, ensure_ascii=False, indent=2)

        logger.info(f"Successfully imported {len(recipes)} recipes (skipped {skipped})")
        logger.info(f"Saved to: {output_path}")
        return True

    except Exception as e:
        logger.error(f"Import failed: {e}")
        return False

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    excel_file = sys.argv[1]
    output_file = None
    merge_mode = False

    # Parse arguments
    for i in range(2, len(sys.argv)):
        if sys.argv[i] == '--output' and i + 1 < len(sys.argv):
            output_file = sys.argv[i + 1]
        elif sys.argv[i] == '--merge':
            merge_mode = True

    success = import_recipes(excel_file, output_file, merge_mode)
    sys.exit(0 if success else 1)
